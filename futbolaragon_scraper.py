import os
import re
import time
import json
import random
from datetime import datetime
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, TimeoutError
from playwright_stealth import Stealth
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

class FutbolAragonScraper:
    def __init__(self, headless=True):
        self.headless = headless
        self.base_url = "https://www.futbolaragon.com"
        
    def setup_browser(self):
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=self.headless,
            args=["--disable-blink-features=AutomationControlled"]
        )
        self.context = self.browser.new_context(
            locale="es-ES",
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        self.page = self.context.new_page()
        Stealth().apply_stealth_sync(self.page)
    
    def teardown_browser(self):
        if hasattr(self, 'browser'):
            self.browser.close()
        if hasattr(self, 'playwright'):
            self.playwright.stop()
            
    def handle_cookie_consent(self):
        print("[INFO] Navigating to home page to bypass cookie consent...")
        self.page.goto(self.base_url)
        try:
            self.page.wait_for_selector(".cmpboxbtnyes", timeout=5000)
            self.page.locator(".cmpboxbtnyes").click()
            print("[SUCCESS] Cookie consent accepted.")
            self.page.wait_for_timeout(1000)
        except TimeoutError:
            print("[INFO] No cookie consent wall detected or already accepted.")
        except Exception as e:
            print(f"[WARNING] Could not click cookie consent: {e}")

    def get_total_jornadas(self, cod_competicion, cod_grupo, cod_delegacion=1, cod_temporada=21):
        url = f"{self.base_url}/pnfg/NPcd/NFG_CmpJornada?cod_primaria=1000120&CodCompeticion={cod_competicion}&CodGrupo={cod_grupo}&CodTemporada={cod_temporada}&CodJornada=1&Sch_Codigo_Delegacion={cod_delegacion}"
        print("[INFO] Discovering total number of jornadas...")
        try:
            self.page.goto(url)
            self.page.wait_for_selector("select#jornada", state="attached", timeout=10000)
            html = self.page.content()
            soup = BeautifulSoup(html, 'html.parser')
            
            select_jornada = soup.find('select', id='jornada')
            if not select_jornada:
                return 0
                
            options = select_jornada.find_all('option')
            # Extract highest option value that is a digit and not 0
            jornadas = [int(opt['value']) for opt in options if opt.get('value') and opt['value'].isdigit() and opt['value'] != '0']
            
            if jornadas:
                total = max(jornadas)
                print(f"[SUCCESS] Discovered {total} jornadas in this competition.")
                return total
            return 0
        except Exception as e:
            print(f"[ERROR] Failed to discover jornadas: {e}")
            self.page.screenshot(path="last_error.png")
            return 0

    def extract_matches_from_competition(self, cod_competicion, cod_grupo, jornadas, target_team, cod_delegacion=1, cod_temporada=21):
        """
        Loops through jornadas and extracts match URLs for a given competition, only for the target team.
        """
        match_urls = []
        for jornada in range(1, jornadas + 1):
            url = f"{self.base_url}/pnfg/NPcd/NFG_CmpJornada?cod_primaria=1000120&CodCompeticion={cod_competicion}&CodGrupo={cod_grupo}&CodTemporada={cod_temporada}&CodJornada={jornada}&Sch_Codigo_Delegacion={cod_delegacion}"
            print(f"[{jornada}/{jornadas}] Fetching Jornada...", end=" ")
            
            try:
                self.page.goto(url)
                self.page.wait_for_selector("table", timeout=10000)
                html = self.page.content()
                soup = BeautifulSoup(html, 'html.parser')
                
                # Each match is usually in a table with width="100%"
                match_tables = soup.find_all('table', width="100%")
                
                added = 0
                for table in match_tables:
                    text_content = table.get_text().lower()
                    if target_team.lower() in text_content:
                        acta_link = table.find('a', href=re.compile(r'NFG_CmpPartido.*CodActa='))
                        
                        # Extract exact Date and Time from the row text (e.g. "28-03-2026 \n 13:15")
                        full_txt = table.get_text(separator=' | ', strip=True)
                        date_match = re.search(r'(\d{2}-\d{2}-\d{4})', full_txt)
                        time_match = re.search(r'(\d{2}:\d{2})', full_txt)
                        fecha_str = date_match.group(1) if date_match else ""
                        hora_str = time_match.group(1) if time_match else ""

                        if acta_link:
                            href = acta_link.get('href')
                            if href:
                                full_url = href if href.startswith('http') else f"{self.base_url}{href}"
                                # Ensure no duplicates
                                if not any(m['url'] == full_url for m in match_urls):
                                    match_urls.append({
                                        'url': full_url,
                                        'jornada': jornada,
                                        'fecha': fecha_str,
                                        'hora': hora_str
                                    })
                                    added += 1
                                    
                if added > 0:
                    print(f"[{jornada}/{jornadas}] Fetching Jornada... Found {added} match(es) for {target_team}.")
                else:
                    print(f"[{jornada}/{jornadas}] Fetching Jornada... No match found.")
                    
                # Add random delay to avoid anti-scraping blocks
                time.sleep(random.uniform(1.5, 3.5))
                
            except Exception as e:
                print(f"Error fetching jornada {jornada}: {e}")
                self.page.screenshot(path="last_error.png")
                # Even on error, wait to let the server recover
                time.sleep(random.uniform(5.0, 10.0))
                
        # Sort chronically to guarantee match sequence in the output DB and Excel
        match_urls.sort(key=lambda x: x['jornada'])
        return match_urls
        
    def parse_match_report(self, match_obj, target_team, match_length=80):
        url = match_obj['url']
        print(f"[INFO] Parsing match: J{match_obj['jornada']} | {url}")
        try:
            self.page.goto(url)
            self.page.wait_for_selector("table", timeout=10000)
            html = self.page.content()
        except Exception as e:
            print(f"[ERROR] Failed to load match report {url}: {e}")
            return None
            
        soup = BeautifulSoup(html, 'html.parser')
        
        data = {
            'match_info': {},
            'teams': {},
            '_all_cards': [],
            '_all_goals': []
        }
        
        # Match Info
        data['match_info']['subtitle'] = f"Jornada {match_obj['jornada']} - {match_obj['fecha']} {match_obj['hora']}"
        data['match_info']['jornada'] = match_obj['jornada']
        data['match_info']['fecha'] = match_obj['fecha']
        data['match_info']['hora'] = match_obj['hora']
        
        div_header = soup.find('h4', style=lambda value: value and 'display: inline-block' in value)
        if div_header:
            data['match_info']['division'] = div_header.get_text(strip=True)

        home_team_elem = soup.find('div', class_='font_widgetL')
        away_team_elem = soup.find('div', class_='font_widgetV')
        
        home_team = home_team_elem.get_text(strip=True) if home_team_elem else "Local"
        away_team = away_team_elem.get_text(strip=True) if away_team_elem else "Visitante"
        data['match_info']['home_team'] = home_team
        data['match_info']['away_team'] = away_team


        # Score Extraction (Improved v7.3)
        marcador = soup.find('div', class_='font_widget_marcador')
        if not marcador:
            # Fallback: find any div that looks like a score (X - Y)
            marcador = soup.find('div', string=re.compile(r'\d+\s*-\s*\d+'))
            
        if marcador:
            score_text = marcador.get_text(strip=True)
            data['match_info']['score'] = score_text
            try:
                # Clean score (sometimes it has weird chars)
                clean_score = re.search(r'(\d+)\s*-\s*(\d+)', score_text)
                if clean_score:
                    h_score = int(clean_score.group(1))
                    a_score = int(clean_score.group(2))
                    
                    target_lower = target_team.lower()
                    # More flexible match: check if target_team is PART of home_team
                    is_home = target_lower in home_team.lower()
                    
                    my_score = h_score if is_home else a_score
                    rival_score = a_score if is_home else h_score
                    
                    if my_score > rival_score: result = "W"
                    elif my_score < rival_score: result = "L"
                    else: result = "D"
                    data['match_info']['result'] = result
            except Exception as e:
                print(f"[DEBUG] Score parse error: {e}")
                data['match_info']['result'] = "Unknown"

        # STRUCTURE-AWARE PARSER (v7)
        # The page structure is: div.details > (div.number + div.desc)
        # div.number = section name (team name, "Goles", "Árbitros", etc.)
        # div.desc = content for that section
        
        for details_block in soup.find_all('div', class_='details'):
            number_div = details_block.find('div', class_='number')
            desc_div = details_block.find('div', class_='desc')
            
            if not number_div or not desc_div:
                continue
            
            section_name = number_div.get_text(strip=True)
            
            # GOALS block: "Goles" is a standalone section
            if 'Goles' in section_name:
                table = desc_div.find('table')
                if table:
                    for row in table.find_all('tr'):
                        cols = row.find_all('td')
                        if len(cols) >= 2:
                            name_text = cols[1].get_text(strip=True)
                            num_goals = max(1, len(re.findall(r"'", name_text)))
                            clean_name = re.sub(r'\(.*?\)', '', name_text).strip()
                            clean_name = re.sub(r'^\d+\-?\d*', '', clean_name).strip()
                            data['_all_goals'].append({
                                'name': clean_name,
                                'count': num_goals
                            })
                continue
            
            # TEAM block: team name contains a dash (e.g. HUESCA-S.D., FRAGA-FÚTBOL BASE)
            if '-' in section_name and 'Árbitros' not in section_name:
                team_name = section_name
                data['teams'][team_name] = {'starters': [], 'subs': [], 'substitutions': [], 'goals': []}
                
                current_section = None
                for child in desc_div.children:
                    if not hasattr(child, 'name') or not child.name:
                        continue
                    
                    child_txt = child.get_text(strip=True)
                    
                    # h5 = section subtitle (Titulares, Suplentes, etc.)
                    if child.name in ['h5', 'h4']:
                        if 'Titulares' in child_txt: current_section = 'starters'
                        elif 'Suplentes' in child_txt: current_section = 'subs'
                        elif 'Sustituciones' in child_txt: current_section = 'substitutions'
                        elif 'Tarjetas' in child_txt: current_section = 'cards'
                        elif 'Cuerpo Técnico' in child_txt: current_section = None
                    
                    elif child.name == 'table' and current_section:
                        if current_section in ['starters', 'subs']:
                            for row in child.find_all('tr'):
                                cols = row.find_all('td')
                                if len(cols) >= 2:
                                    num = cols[0].get_text(strip=True)
                                    if not num.isdigit(): continue
                                    name = cols[1].get_text(strip=True)
                                    data['teams'][team_name][current_section].append({'number': num, 'name': name, 'yellow_cards': 0, 'red_cards': 0})
                        
                        elif current_section == 'cards':
                            for row in child.find_all('tr'):
                                cols = row.find_all('td')
                                if len(cols) >= 2:
                                    card_info = cols[1].get_text(strip=True)
                                    # Extract name using regex (handling minutes like (80'))
                                    name_match = re.search(r'^(.*?)\s*\((\d+)\'?\)', card_info)
                                    clean_name = name_match.group(1).strip() if name_match else card_info.strip()
                                    
                                    img_tag = cols[0].find('img')
                                    src = str(img_tag.get('src', '')).lower() if img_tag else ""
                                    
                                    card_type = None
                                    if 'amarilla' in src or 'yellow' in src: card_type = 'Y'
                                    elif 'roja' in src or 'red' in src or 'rojo' in src: card_type = 'R'
                                    
                                    if card_type:
                                        print(f"  [DEBUG] Card found: {card_type} for {clean_name}")
                                        data['_all_cards'].append({'name': clean_name, 'type': card_type})
                        
                        elif current_section == 'substitutions':
                            rows = child.find_all('tr')
                            if len(rows) >= 2:
                                in_row, out_row = rows[0], rows[1]
                                in_cols, out_cols = in_row.find_all('td'), out_row.find_all('td')
                                
                                in_num = in_cols[0].get_text(strip=True) if in_cols else '0'
                                out_num = out_cols[0].get_text(strip=True) if out_cols else '0'
                                out_txt = out_cols[1].get_text(strip=True) if len(out_cols) > 1 else ''
                                
                                min_match = re.search(r"\((\d+)'?\)", out_txt)
                                minute = int(min_match.group(1)) if min_match else 0
                                
                                data['teams'][team_name]['substitutions'].append({
                                    'in_number': in_num,
                                    'out_number': out_num,
                                    'minute': minute
                                })
        
        # Now assign goals to the correct team using fuzzy word matching
        all_goals = data.get('_all_goals', [])
        if all_goals and data['teams']:
            for team_name, team_data_inner in data['teams'].items():
                all_players = {p['name'] for p in team_data_inner['starters'] + team_data_inner['subs']}
                for g in all_goals:
                    g_words = set(g['name'].lower().replace(',', '').split())
                    for p_name in all_players:
                        p_words = set(p_name.lower().replace(',', '').split())
                        if len(g_words.intersection(p_words)) >= 1:
                            team_data_inner['goals'].append(g)
                            break
            del data['_all_goals']
            
        # Assign cards using fuzzy matching
        all_cards = data.get('_all_cards', [])
        if all_cards and data['teams']:
            for team_name, team_data_inner in data['teams'].items():
                target_players_list = team_data_inner['starters'] + team_data_inner['subs']
                for c in all_cards:
                    c_name_norm = c['name'].lower().replace(',', '').strip()
                    c_words = set(c_name_norm.split())
                    
                    for p in target_players_list:
                        p_name_norm = p['name'].lower().replace(',', '').strip()
                        p_words = set(p_name_norm.split())
                        
                        if c_name_norm == p_name_norm or c_words.issubset(p_words) or p_words.issubset(c_words):
                            if c['type'] == 'Y': p['yellow_cards'] = p.get('yellow_cards', 0) + 1
                            else: p['red_cards'] = p.get('red_cards', 0) + 1
                            print(f"  [DEBUG] Assigned {c['type']} card to {p['name']}")
                            break
            del data['_all_cards']
        
        # Proceed to calculate stats
        target_stats = self._calculate_player_stats(data, target_team, match_length)
        if not target_stats:
            for t_name in data['teams'].keys():
                if target_team.lower() in t_name.lower():
                    target_stats = self._calculate_player_stats(data, t_name, match_length)
                    break
                    
        return {
            'match_info': data['match_info'],
            'player_stats': target_stats
        }
        
    def _calculate_player_stats(self, parsed_data, target_team, match_length):
        team_data = parsed_data['teams'].get(target_team)
        if not team_data: return []

        player_stats = {}
        for p in team_data['starters']:
            player_stats[p['name']] = {
                'number': p['number'],
                'is_starter': True,
                'minutes_played': match_length,
                'entry_minute': 0,
                'exit_minute': match_length,
                'goals': 0,
                'yellow_cards': p.get('yellow_cards', 0),
                'red_cards': p.get('red_cards', 0)
            }
        
        for p in team_data['subs']:
            player_stats[p['name']] = {
                'number': p['number'],
                'is_starter': False,
                'minutes_played': 0,
                'entry_minute': None,
                'exit_minute': None,
                'goals': 0,
                'yellow_cards': p.get('yellow_cards', 0),
                'red_cards': p.get('red_cards', 0)
            }

        number_to_name = {stats['number']: name for name, stats in player_stats.items()}

        for sub in team_data['substitutions']:
            minute = sub.get('minute') or match_length
            out_name = number_to_name.get(sub.get('out_number', ''))
            out_p = player_stats.get(out_name) if out_name else None
            if out_p:
                out_p['exit_minute'] = minute
                if out_p['entry_minute'] is not None:
                    out_p['minutes_played'] = minute - out_p['entry_minute']
            
            in_name = number_to_name.get(sub.get('in_number', ''))
            in_p = player_stats.get(in_name) if in_name else None
            if in_p:
                in_p['entry_minute'] = minute
                in_p['exit_minute'] = match_length
                in_p['minutes_played'] = match_length - minute

        for g in team_data.get('goals', []):
            g_words = set(g['name'].lower().replace(',', '').split())
            if not g_words: continue
            
            best_match, best_score = None, 0
            for p_name, stats in player_stats.items():
                p_words = set(p_name.lower().replace(',', '').split())
                score = len(g_words.intersection(p_words))
                if score > best_score:
                    best_score, best_match = score, stats
            
            if best_match and best_score > 0:
                best_match['goals'] += g['count']

        stats_list = []
        for name, stats in player_stats.items():
            stats_list.append({'name': name, **stats})
        
        return sorted(stats_list, key=lambda x: (-x['minutes_played'], x['name']))

class ExcelExporter:
    def __init__(self, filename="futbolaragon_data.xlsx", target_team="HUESCA-S.D.", target_player=None):
        self.filename = filename
        self.target_team = target_team
        self.target_player = target_player
        self.matches_data = []
        self.players_data = []
        import pandas as pd
        self.pd = pd
        
    def append_match_data(self, match_data):
        info = match_data['match_info']
        match_title = f"{info.get('home_team')} vs {info.get('away_team')}"
        home, away = info.get('home_team', ''), info.get('away_team', '')
        
        condicion = "Local" if self.target_team.lower() in home.lower() else "Visitante"
        rival = away if condicion == "Local" else home
        
        self.matches_data.append({
            "Jornada/Fecha": info.get('subtitle', ''),
            "División": info.get('division', ''),
            "Partido": match_title,
            "Equipo": self.target_team,
            "Rival": rival,
            "Condición": condicion,
            "Score": info.get('score', ''),
            "Result": info.get('result', '')
        })
        
        for p in match_data.get('player_stats', []):
            titular = "Sí" if p['is_starter'] else "No"
            min_played = p['minutes_played']
            self.players_data.append({
                "Partido": match_title,
                "Rival": rival,
                "Condición": condicion,
                "Jugador": p['name'],
                "Dorsal": p['number'],
                "Titular": titular,
                "Min. Entrada": p['entry_minute'] if p['entry_minute'] is not None else 0,
                "Min. Salida": p['exit_minute'] if p['exit_minute'] is not None else 0,
                "Minutos Jugados": min_played,
                "Goles": p.get('goals', 0),
                "Amarillas": p.get('yellow_cards', 0),
                "Rojas": p.get('red_cards', 0),
                "Jugó": "Sí" if min_played > 0 else "No",
                "Suplente Usado": "Sí" if titular == "No" and min_played > 0 else "No",
                "Titularidad_num": 1 if p['is_starter'] else 0
            })
            
    def save(self):
        pd = self.pd
        df_matches, df_players = pd.DataFrame(self.matches_data), pd.DataFrame(self.players_data)
        
        df_resumen = pd.DataFrame()
        if not df_players.empty:
            resumen = df_players.groupby("Jugador").agg(
                Partidos_Convocado=("Partido", "count"),
                Titularidades=("Titularidad_num", "sum"),
                Minutos_Totales=("Minutos Jugados", "sum"),
                Goles_Totales=("Goles", "sum"),
                Amarillas=("Amarillas", "sum"),
                Rojas=("Rojas", "sum")
            ).reset_index()
            
            suplencias_usado = df_players[df_players["Suplente Usado"] == "Sí"].groupby("Jugador").size().reset_index(name="Suplencias Usado")
            partidos_jugados = df_players[df_players["Jugó"] == "Sí"].groupby("Jugador").size().reset_index(name="Partidos Jugados")
            
            resumen = resumen.merge(suplencias_usado, on="Jugador", how="left").fillna({"Suplencias Usado": 0})
            resumen = resumen.merge(partidos_jugados, on="Jugador", how="left").fillna({"Partidos Jugados": 0})
            
            for col in ["Suplencias Usado", "Partidos Jugados", "Titularidades", "Amarillas", "Rojas"]:
                resumen[col] = resumen[col].astype(int)
            
            resumen["Media Min/PJ"] = (resumen["Minutos_Totales"] / resumen.replace({'Partidos Jugados': {0: pd.NA}})["Partidos Jugados"]).fillna(0).round(1)
            resumen["% Titularidad"] = ((resumen["Titularidades"] / resumen.replace({'Partidos Jugados': {0: pd.NA}})["Partidos Jugados"]) * 100).fillna(0).round(1)
            df_resumen = resumen.sort_values("Minutos_Totales", ascending=False)
            
        json_path = "dashboard/public/futbolaragon_data.json"
        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump({"matches": self.matches_data, "players": self.players_data}, f, ensure_ascii=False, indent=2)
            
        print(f"[SUCCESS] Web dashboard data exported to {json_path}")
        
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            if not df_matches.empty: df_matches.to_excel(writer, sheet_name="Detalle Partidos", index=False)
            if not df_players.empty: df_players.to_excel(writer, sheet_name="Datos Jugadores", index=False)
            if not df_resumen.empty: df_resumen.to_excel(writer, sheet_name="Resumen Plantilla", index=False)
        
        from openpyxl import load_workbook
        wb = load_workbook(self.filename)
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = max((len(str(cell.value)) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2
            if ws.dimensions: ws.auto_filter.ref = ws.dimensions
        wb.save(self.filename)

def main():
    TARGET_TEAM, COD_COMPETICION, COD_GRUPO = "HUESCA-S.D.", "22320178", "22401727"
    scraper = FutbolAragonScraper(headless=True)
    exporter = ExcelExporter(filename="futbolaragon_data.xlsx", target_team=TARGET_TEAM)
    
    try:
        scraper.setup_browser()
        scraper.handle_cookie_consent()
        total_jornadas = scraper.get_total_jornadas(COD_COMPETICION, COD_GRUPO) or 1
        match_urls = scraper.extract_matches_from_competition(COD_COMPETICION, COD_GRUPO, total_jornadas, TARGET_TEAM)
        
        for i, m_obj in enumerate(match_urls, 1):
            print(f"--- Processing match {i}/{len(match_urls)} ---")
            m_data = scraper.parse_match_report(m_obj, TARGET_TEAM)
            if m_data: exporter.append_match_data(m_data)
            if i < len(match_urls): time.sleep(random.uniform(1.5, 3.5))
            
        exporter.save()
        print("\n[GIT] Pushing updated data to GitHub...")
        import subprocess
        try:
            subprocess.run(["git", "add", "dashboard/public/futbolaragon_data.json"], check=True)
            subprocess.run(["git", "commit", "-m", "chore: update match data [scraper]"], check=True)
            subprocess.run(["git", "push"], check=True)
            print("[GIT] ✅ Data pushed.")
        except: pass
    finally: scraper.teardown_browser()

if __name__ == "__main__": main()
